import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()

# ============ Sheet 1: Chinese Version ============
ws_cn = wb.active
ws_cn.title = "相关性分析总结_中文"

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
subheader_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
linear_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
nonlinear_fill = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Title
ws_cn.merge_cells('A1:F1')
ws_cn['A1'] = "健康与运动科学领域 - 相关性分析类型总结"
ws_cn['A1'].font = Font(bold=True, size=14)
ws_cn['A1'].alignment = Alignment(horizontal="center")

# Headers
headers_cn = ["类别", "关系类型", "描述", "健康领域示例", "检测技术", "Python实现"]
for col, header in enumerate(headers_cn, 1):
    cell = ws_cn.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# Data rows - Chinese
data_cn = [
    ["线性", "正相关", "变量A增加，变量B也增加，呈均匀比例", "步数↑ → 卡路里消耗↑\n运动时长 ↔ 心率升高\n训练负荷 ↔ 疲劳感", "Pearson相关系数", "scipy.stats.pearsonr()"],
    ["线性", "负相关", "变量A增加，变量B减少，呈均匀比例", "年龄↑ → 最大心率↓\n压力指数 ↔ HRV\n久坐时间 ↔ 心肺适能", "Pearson相关系数", "scipy.stats.pearsonr()"],
    ["非线性", "单调非线性", "持续上升或下降，但变化速率不均匀（边际效应递减）", "运动频率↑ → 心肺适能↑（0→3次/周提升显著，3→6次提升变小）\n冥想时长 ↔ 压力降低（前10分钟效果最大）", "Spearman秩相关", "scipy.stats.spearmanr()"],
    ["非线性", "U型关系", "存在最低点，两端值都较高", "BMI ↔ 死亡率（过轻和过重都增加风险）\n钠摄入量 ↔ 心血管风险", "多项式回归（二次项）", "numpy.polyfit(x, y, 2)"],
    ["非线性", "倒U型关系", "存在最优点，两端值都较低", "运动量 ↔ HRV（适量最佳）\n睡眠时长 ↔ 认知表现（7-8小时最佳）\n咖啡因 ↔ 运动表现", "多项式回归（二次项系数为负）", "numpy.polyfit(x, y, 2)"],
    ["非线性", "阈值效应", "在某个临界点发生突变", "血氧SpO2：>95%正常，<90%危险\n睡眠：<5小时认知急剧下降\n心率：超过最大心率85%进入无氧区", "分段回归", "pwlf库 / 自定义分段模型"],
    ["非线性", "滞后效应", "影响延迟一定时间后出现", "今日运动 → 明日HRV变化（24-48小时）\n压力事件 → 睡眠质量（1-2天后影响）\n饮食变化 → 体重（数周）", "时间滞后相关分析", "pandas.shift() + 相关分析"],
    ["非线性", "交互效应", "多变量组合产生特殊效果，单独可能无关联", "咖啡因+晚间运动 → 睡眠差\n压力+酒精 → HRV影响更大\n海拔+运动强度 → 心率反应异常", "互信息 / 随机森林", "sklearn.mutual_info_regression()\nsklearn.RandomForestRegressor()"],
]

for row_idx, row_data in enumerate(data_cn, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_cn.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = left_align if col_idx > 2 else center_align
        cell.border = thin_border
        if row_data[0] == "线性":
            cell.fill = linear_fill
        else:
            cell.fill = nonlinear_fill

# Priority table
ws_cn.cell(row=14, column=1, value="MVP实现优先级").font = Font(bold=True, size=12)
priority_headers_cn = ["优先级", "关系类型", "理由"]
for col, header in enumerate(priority_headers_cn, 1):
    cell = ws_cn.cell(row=15, column=col, value=header)
    cell.font = header_font
    cell.fill = subheader_fill
    cell.alignment = center_align
    cell.border = thin_border

priority_data_cn = [
    ["P0 (必须)", "线性相关 + 单调非线性 + 倒U型", "基础功能，低成本，核心场景"],
    ["P1 (重要)", "阈值效应 + 滞后效应", "健康警示，行为洞察"],
    ["P2 (增强)", "交互效应", "自动发现复杂模式"],
]

for row_idx, row_data in enumerate(priority_data_cn, 16):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_cn.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = center_align if col_idx == 1 else left_align
        cell.border = thin_border

# Column widths for Chinese sheet
col_widths_cn = [10, 14, 35, 45, 25, 35]
for i, width in enumerate(col_widths_cn, 1):
    ws_cn.column_dimensions[get_column_letter(i)].width = width

# Row heights
for row in range(4, 12):
    ws_cn.row_dimensions[row].height = 60

# ============ Sheet 2: English Version ============
ws_en = wb.create_sheet("Correlation_Analysis_EN")

# Title
ws_en.merge_cells('A1:F1')
ws_en['A1'] = "Health & Sports Science - Correlation Analysis Types Summary"
ws_en['A1'].font = Font(bold=True, size=14)
ws_en['A1'].alignment = Alignment(horizontal="center")

# Headers
headers_en = ["Category", "Relationship Type", "Description", "Health Domain Examples", "Detection Technique", "Python Implementation"]
for col, header in enumerate(headers_en, 1):
    cell = ws_en.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# Data rows - English
data_en = [
    ["Linear", "Positive Correlation", "Variable A increases, Variable B also increases proportionally", "Steps↑ → Calories burned↑\nExercise duration ↔ Heart rate elevation\nTraining load ↔ Fatigue perception", "Pearson Correlation", "scipy.stats.pearsonr()"],
    ["Linear", "Negative Correlation", "Variable A increases, Variable B decreases proportionally", "Age↑ → Max heart rate↓\nStress index ↔ HRV\nSedentary time ↔ Cardio fitness", "Pearson Correlation", "scipy.stats.pearsonr()"],
    ["Non-linear", "Monotonic Non-linear", "Continuous increase/decrease at non-uniform rate (diminishing returns)", "Exercise frequency↑ → Cardio fitness↑ (0→3x/week significant, 3→6x smaller gains)\nMeditation ↔ Stress reduction (first 10 min most effective)", "Spearman Rank Correlation", "scipy.stats.spearmanr()"],
    ["Non-linear", "U-shaped", "Minimum point exists, higher values at both ends", "BMI ↔ Mortality (both underweight & overweight increase risk)\nSodium intake ↔ Cardiovascular risk", "Polynomial Regression (quadratic)", "numpy.polyfit(x, y, 2)"],
    ["Non-linear", "Inverted U-shaped", "Optimal point exists, lower values at both ends", "Exercise volume ↔ HRV (moderate is best)\nSleep duration ↔ Cognitive performance (7-8h optimal)\nCaffeine ↔ Exercise performance", "Polynomial Regression (negative quadratic)", "numpy.polyfit(x, y, 2)"],
    ["Non-linear", "Threshold Effect", "Sudden change at a critical point", "SpO2: >95% normal, <90% dangerous\nSleep: <5h causes sharp cognitive decline\nHeart rate: >85% max HR enters anaerobic zone", "Piecewise Regression", "pwlf library / custom segmented model"],
    ["Non-linear", "Lag Effect", "Impact appears with time delay", "Today's exercise → Tomorrow's HRV change (24-48h)\nStress event → Sleep quality (1-2 days later)\nDietary changes → Weight (weeks)", "Time-lagged Correlation", "pandas.shift() + correlation"],
    ["Non-linear", "Interaction Effect", "Multiple variables combined produce special effects; individually may show no correlation", "Caffeine + Evening exercise → Poor sleep\nStress + Alcohol → Greater HRV impact\nAltitude + Exercise intensity → Abnormal HR response", "Mutual Information / Random Forest", "sklearn.mutual_info_regression()\nsklearn.RandomForestRegressor()"],
]

for row_idx, row_data in enumerate(data_en, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_en.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = left_align if col_idx > 2 else center_align
        cell.border = thin_border
        if row_data[0] == "Linear":
            cell.fill = linear_fill
        else:
            cell.fill = nonlinear_fill

# Priority table - English
ws_en.cell(row=14, column=1, value="MVP Implementation Priority").font = Font(bold=True, size=12)
priority_headers_en = ["Priority", "Relationship Type", "Rationale"]
for col, header in enumerate(priority_headers_en, 1):
    cell = ws_en.cell(row=15, column=col, value=header)
    cell.font = header_font
    cell.fill = subheader_fill
    cell.alignment = center_align
    cell.border = thin_border

priority_data_en = [
    ["P0 (Must Have)", "Linear + Monotonic Non-linear + Inverted U", "Foundational, low cost, core scenarios"],
    ["P1 (Important)", "Threshold Effect + Lag Effect", "Health alerts, behavioral insights"],
    ["P2 (Enhancement)", "Interaction Effect", "Auto-discover complex patterns"],
]

for row_idx, row_data in enumerate(priority_data_en, 16):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_en.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = center_align if col_idx == 1 else left_align
        cell.border = thin_border

# Column widths for English sheet
col_widths_en = [12, 18, 40, 50, 28, 35]
for i, width in enumerate(col_widths_en, 1):
    ws_en.column_dimensions[get_column_letter(i)].width = width

# Row heights
for row in range(4, 12):
    ws_en.row_dimensions[row].height = 60

# Save workbook
wb.save(r'd:\Antigravity WS\H&F Lib\docs\correlation_analysis_summary.xlsx')
print("Excel file created successfully!")
